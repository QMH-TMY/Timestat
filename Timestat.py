#!/usr/bin/python3
# -*- coding: utf-8 -*-
#
#    Author: Shieber
#
#                             APACHE LICENSE
#    Licensed under the Apache License, Version 2.0 (the "License"); you may
#    not use this file except in compliance with the License. You may obtain
#    a copy of the License at http://www.apache.org/licenses/LICENSE-2.0
#    Unless required by applicable law or agreed to in writing, software
#    distributed under the License is distributed on an "AS IS" BASIS, WITHOUT
#    WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the
#    License for the specific language governing permissions and limitations
#    under the License.
#
#                            Function Description
#    1.通过邮件统计每天花费的时间
#    2.表保存的格式为xlsx         
#    3.后期将加入每月总花费时间统计的功能      
#
#    Copyright 2019 
#    All Rights Reserved!


import poplib 
import logging
import requests
import openpyxl
import datetime
import os,re,time

from email.parser import Parser 
from email.utils import parseaddr 
from email.header import decode_header 

class TimeManage():
    def __init__(self):
        '''
           初始化邮件和excel表的位置，
           所有的值依据自己的电脑自行设定
        '''
        self.sleep_time = 2                             #写入excel的间隔时间，
        self.year_key   = 'year'                        #设立新表的依据
        self.all_sheet  = 'All'                         #设置总表的名称 
        self.inbox_dir  = '/home/shieber/files/gitproject/timestat/inbox.txt'  #暂存位置
        self.excel_dir  = '/home/shieber/files/gitproject/timestat/'     #excel存放路径 
        self.basename   = "timestat.xlsx"               #excel表基本名
        self.date_str   = 'date'
        self.sheet_info = {                             #excel表项名称,时间，消费方式等等
                            "1":'Date(y-m-d)',		#日期时间 
                            "2":'Method',               #做的事
                            "3":'Time(h:m)',            #花费时间xhxm
                            "4":'Notes'                 #注
                           }
        self.detail_tm  = self.date_time()              #设置时间

    ########1.记录统计日期时间#########################
    def date_time(self):
        '''
                返回记录的日期，格式为：2019-01-10
                该值是写入excel表第一列的数据
        '''
        date = datetime.datetime.now()
        year = str(date.year)
        mon  = str(date.month)
        day  = str(date.day)
        date_s = year + "-" + mon + "-" + day
        return date_s

        ########2.提取邮件中的消费信息######################
    def get_time_info(self):
        '''从邮件中提取所需的消费内容'''
        emailObj = open(self.inbox_dir)
        textline = emailObj.readline()
        date_p = re.compile(r'\((\d+)(-|/|\.|\s)(\d+)(-|/|\.|\s)(\d+)\)') #正则查找日期
        date   = date_p.findall(textline)
        if date:
            year   = str(date[0][0]) 
            time_date = "".join(date[0])
        else:
            year   = str(datetime.datetime.now().year)                 #使用当前时间
            time_date = self.detail_tm
                
        textline = emailObj.readline()
        time_p = re.compile(r'(\w+?)(:|\s)?(\d+)(H|h|M|m)(\d+)?(m)?')  #正则查找时间项
        time   = time_p.findall(textline)
        if time:
            time_info = self.distil_time_dic(year,time_date, time)
        else:
            time_info = {}                                             #没有信息

        emailObj.close()
        return time_info 

    def distil_time_dic(self, year, time_date, time):
        '''解析时间内容成字典格式并返回
           time格式为[('jd',':','1','h','21','m'),(),()]
        '''
        time_info_d  = {}
        time_info_d[self.year_key] = year
        time_info_d[self.date_str] = time_date 
        for t in range(len(time)):
            tuple_t = time[t]                    #解析时间数据为元组
            method  = tuple_t[0]                 #解析做的事件
            hours   = tuple_t[2:]                #解析出消耗的时间
            hm_s = "".join(hours)
            time_info_d[method] = hm_s 
        return time_info_d 

        ########4.向Excel中写入消费信息################
    def write_to_excel(self):
        '''向20xxtimestat.xlsx表中写入所有信息(核心函数)'''
        time_info = self.get_time_info()
        if not time_info:
            sys.exit(-1)                         #没有时间信息则直接退出

        excel_name = self.excel_dir + str(time_info[self.year_key]) + self.basename 

        ########关键点，易出错#####################
        keys   = time_info.keys()
        del keys[keys.index(self.year_key)]
        del keys[keys.index(self.date_str)]

        if not os.path.exists(excel_name):       #判断对应年文件是否存在,不存在就创建
            keys.append(self.all_sheet)      #加入总表All
            self.create_year_sheet(keys, excel_name)
            del keys[keys.index(self.all_sheet)]

        wb = openpyxl.load_workbook(excel_name)  #打开对应年的文件写入信息
        sheets = wb.get_sheet_names()
        for key in keys:
            key_title = key.title()              #事件的首字符大写
            if key_title not in sheets:
                wb.create_sheet(1,key_title)     #为新的事情添加分表
                self.add_item(wb,key_title)

            sheet_lis = [self.all_sheet, key_title]         #每笔时间记录到总表和分表中
            for sheet in sheet_lis:
                curren_s = wb.get_sheet_by_name(sheet)      #开总表和分表记录时间
                n_row  = str(curren_s.get_highest_row() + 1)#设置写入的行数:最大行加1
                curren_s['A' + n_row] = time_info[self.date_str]   #日期
                curren_s['B' + n_row] = key                 #事情
                curren_s['C' + n_row] = time_info[key]      #时间

        wb.save(excel_name)
        time.sleep(self.sleep_time)              #稍停顿，待excel表数据存储完毕

        ########5.创建数据记录表20xxdebt.xlsx##########
    def create_year_sheet(self,sheets, excel_name):
        '''如果不存在某年的表就建立相应的表(20xxdebt.xlsx)'''
        wb = openpyxl.Workbook()
        for sheet in sheets:
            wb.create_sheet(0, sheet.title())    #首字母大写
                                                 
        sheet_names = wb.get_sheet_names()
        for sheet_name in sheet_names:
            self.add_item(wb,sheet_name) 
        wb.remove_sheet(wb.get_sheet_by_name('Sheet')) #删除多余的表	
        wb.save(excel_name)

    def add_item(self,wb,name):
        '''为每个表添加时间消费项的标题'''
        sheet = wb.get_sheet_by_name(name)
        sheet['A1'] = self.sheet_info['1']        #初始化表的记录项的相关信息
        sheet['B1'] = self.sheet_info['2']
        sheet['C1'] = self.sheet_info['3']
        sheet['D1'] = self.sheet_info['4']


class EmailManage():
'''继承父类，连接网络查询，下载，调用父类函数写入excel,删除邮箱对应邮件'''
    def __init__(self):
        '''初始化设置相关信息'''
        self.timemanage  =  TimeManage()
        self.email_num   = 10                    #默认读取邮件数量
        self.inbox_dir   = '/home/shieber/timestat/inbox.txt' 
        self.pop3_server = 'pop.163.com'         #对应邮箱的pop3服务器
        self.email_addr  = ""                    #你的邮箱账号
        self.password    = ""                    #你的登录密码
        self.identifier  = 'Time used'           #你发送的邮件标题中的关键字

    def guess_charset(self, msg): 
        '''获取邮件字符集编码'''
        charset = msg.get_charset() 
        if charset is None: 
            content_type = msg.get('Content-Type', '').lower()  #获取失败时再次获取
            pos = content_type.find('charset=') 
            if pos >= 0: 
                charset = content_type[pos + 8:].strip() 
        return charset 

    def print_info(self, msg, indent=0): 
        '''打印出邮件信息'''
        header = "Subject"
        if indent == 0: 
            value = msg.get(header, '') 
            if value: 
                print('%s: %s' % (header, value)) 
           else:
                sys.exit(-1)

        if (msg.is_multipart()): 
            parts = msg.get_payload() 
            for n, part in enumerate(parts):
                self.print_info(part, indent + 1) 
        else: 
            content_type = msg.get_content_type() 
            if content_type=='text/plain' or content_type=='text/html': 
                content = msg.get_payload(decode=True) 
                charset = self.guess_charset(msg) 
                if charset: 
                    content = content.decode(charset) 
                print('Text: %s' % (content)) 

    def write_to_inbox(self, msg):
        '''写入inbox.txt'''
        inboxObj = open(self.inbox_dir,'w')      #每次打开都新建，所以不用管以前的信息
        stdotput = sys.stdout                    #暂时将标准输出存储起来，以便恢复   
        sys.stdout = inboxObj                    #设置系统打印输出到文件
        self.print_info(msg)                     #输出邮件信息到本地文件inbox.txt
        sys.stdout = stdotput                    #恢复系统输出到终端
        inboxObj.close()                         #关闭文件

    def download_write(self, index, server):
        '''下载邮件并写入excel表'''
        try:
            resp, lines, octets = server.retr(index)      #下载邮件
        except Exception as err:
            sys.exit(-1)

        msg_content = b'\r\n'.join(lines).decode('utf-8') #拼接邮件内容 
        msg = Parser().parsestr(msg_content)              #解析邮件内容

        value = msg.get("Subject", '').lower()
        if self.identifier in value:             #标题含有关键的标识符时执行
            self.write_to_inbox(msg)             #先写入暂时文件inbox.txt
            self.timemanage.write_to_excel()     #写入excel(核心函数)
            server.dele(index)                   #删除邮件,核心函数,千万不要误删

    def download_write_delete(self, indexs, server):
        '''下载写入并删除邮件(核心函数)'''
        if indexs > self.email_num:
            for index in range(indexs, indexs - self.email_num, -1):
                self.download_write(index, server)#具体写入函数
        else:
            for index in range(indexs,0,-1):      #不太可能执行,除非邮件数量少于10
                self.download_write(index, server) 
        server.quit()                             

    def connect(self):
        '''连接邮箱服务器'''
        try:
            server = poplib.POP3_SSL(self.pop3_server, 995) 
            server.user(self.email_addr) 
            server.pass_(self.password) 
        except Exception as err:
                return None, None
        resp, mails, octets = server.list() 
        indexs = len(mails) 
        return indexs, server

def stat_main():
    '''主函数:程序入口，直接调用'''
    email_manage = EmailManage()
    indexs, server = email_manage.connect()
    email_manage.download_write_delete(indexs, server)

if __name__ == "__main__":
    stat_main()
